"""Parse 2026-1전화번호부.xlsx → data/phonebook-2026-1.json"""
from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "2026-1전화번호부.xlsx"
OUT = ROOT / "data" / "phonebook-2026-1.json"


def cell_fill_rgb(cell) -> str | None:
    try:
        fg = cell.fill.fgColor if cell.fill else None
        if not fg:
            return None
        rgb = getattr(fg, "rgb", None)
        if rgb and isinstance(rgb, str) and rgb not in ("00000000",):
            if rgb.startswith("Values"):
                return None
            return rgb.upper()
    except Exception:
        return None
    return None


def is_colored(cell) -> bool:
    rgb = cell_fill_rgb(cell)
    if not rgb:
        return False
    return rgb not in ("00000000", "FFFFFFFF", "FF000000")


def clean_unit(s: object) -> str:
    text = str(s or "").strip()
    text = re.sub(r"\s*\([^)]*\)\s*$", "", text)
    return text.strip()


def is_hangul_name(s: object) -> bool:
    text = str(s or "").strip()
    if not text or text.upper() == "FAX":
        return False
    if re.fullmatch(r"[\d\s,\-]+", text):
        return False
    return any("\uac00" <= ch <= "\ud7a3" for ch in text)


def valid_ext(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or "," in s or "-" in s:
        return None
    digits = re.sub(r"[^0-9]", "", s)
    if not digits or len(digits) < 3:
        return None
    return digits


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active

    cur_parent: str | None = None
    cur_unit: str | None = None
    unit_is_target: dict[tuple[str, str], bool] = {}
    unit_publish_names: dict[tuple[str, str], set[str]] = {}
    people: list[dict] = []
    parents: list[dict] = []
    units: list[dict] = []

    for r in range(1, ws.max_row + 1):
        a = ws[f"A{r}"]
        c = ws[f"C{r}"]
        d = ws[f"D{r}"]
        e = ws[f"E{r}"]
        f = ws[f"F{r}"]
        av = a.value
        cv = c.value
        dv = d.value
        ev = str(e.value).strip() if e.value else ""
        fv = str(f.value).strip() if f.value else ""
        bold = bool(a.font and a.font.bold)
        und = bool(a.font and a.font.underline and a.font.underline != "none")
        colored = is_colored(a)

        if av and bold and und and colored:
            cur_parent = clean_unit(av)
            cur_unit = None
            parents.append({"row": r, "name": cur_parent, "raw": str(av)})
            # 상위부서 행에 대상부서 표시 → 상위부서 자체도 선택 옵션
            if cur_parent and ev == "대상부서":
                unit_is_target[(cur_parent, cur_parent)] = True
            if cur_parent and fv and "전체일정" not in fv and is_hangul_name(fv):
                unit_publish_names.setdefault((cur_parent, cur_parent), set()).add(
                    fv.strip()
                )
            continue

        if av and bold and und and not colored:
            cur_unit = clean_unit(av)
            units.append(
                {"row": r, "parent": cur_parent, "name": cur_unit, "raw": str(av)}
            )
            if cur_parent and cur_unit:
                key = (cur_parent, cur_unit)
                if ev == "대상부서":
                    unit_is_target[key] = True
                if fv and "전체일정" not in fv and is_hangul_name(fv):
                    unit_publish_names.setdefault(key, set()).add(fv.strip())
            continue

        name = str(cv).strip() if cv else ""
        ext = valid_ext(dv)
        if is_hangul_name(name) and ext and cur_parent:
            unit = cur_unit or cur_parent
            key = (cur_parent, unit)
            can_publish = bool(fv and "전체일정" in fv)
            if fv and is_hangul_name(fv) and fv.strip() == name:
                can_publish = True
            duty = str(av).strip() if av and not (bold and und) else None
            people.append(
                {
                    "phoneParent": cur_parent,
                    "phoneDept": unit,
                    "name": name,
                    "phoneExt": ext,
                    "duty": duty,
                    "canPublish": can_publish,
                    "row": r,
                }
            )
            if ev == "대상부서":
                unit_is_target[key] = True
            if fv and "전체일정" not in fv and is_hangul_name(fv):
                unit_publish_names.setdefault(key, set()).add(fv.strip())

    for p in people:
        key = (p["phoneParent"], p["phoneDept"])
        if p["name"] in unit_publish_names.get(key, set()):
            p["canPublish"] = True

    target_depts: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for u in units:
        if not u["parent"] or not u["name"]:
            continue
        key = (u["parent"], u["name"])
        if unit_is_target.get(key) and key not in seen:
            seen.add(key)
            target_depts.append(
                {
                    "phoneParent": u["parent"],
                    "phoneDept": u["name"],
                    "label": f"{u['name']}({u['parent']})",
                }
            )
    for key, ok in unit_is_target.items():
        if ok and key not in seen and key[0] and key[1]:
            seen.add(key)
            target_depts.append(
                {
                    "phoneParent": key[0],
                    "phoneDept": key[1],
                    "label": f"{key[1]}({key[0]})",
                }
            )
    target_depts.sort(key=lambda x: (x["phoneParent"], x["phoneDept"]))

    uniq: dict[tuple, dict] = {}
    for p in people:
        k = (p["phoneParent"], p["phoneDept"], p["name"], p["phoneExt"])
        if k in uniq:
            uniq[k]["canPublish"] = uniq[k]["canPublish"] or p["canPublish"]
        else:
            uniq[k] = p
    people_list = list(uniq.values())

    accounts = []
    for p in sorted(
        people_list,
        key=lambda x: (
            x["phoneParent"],
            x["phoneDept"],
            x["name"],
            x["phoneExt"],
        ),
    ):
        base = "|".join(
            [p["phoneParent"], p["phoneDept"], p["name"], p["phoneExt"]]
        )
        hid = hashlib.sha1(base.encode("utf-8")).hexdigest()[:10]
        accounts.append(
            {
                "employeeId": f"pb_{hid}",
                "name": p["name"],
                "department": p["phoneDept"],
                "phoneParent": p["phoneParent"],
                "phoneDept": p["phoneDept"],
                "phoneExt": p["phoneExt"],
                "duty": p.get("duty"),
                "canPublishToOverall": p["canPublish"],
                "role": "USER",
                "isTeamLeader": p["canPublish"],
            }
        )

    out = {
        "source": "2026-1전화번호부.xlsx",
        "targetDepartments": target_depts,
        "accounts": accounts,
        "stats": {
            "parents": len(parents),
            "units": len(units),
            "targetDepartments": len(target_depts),
            "accounts": len(accounts),
            "publishers": sum(1 for a in accounts if a["canPublishToOverall"]),
        },
    }
    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(out["stats"], ensure_ascii=False, indent=2))
    print("--- targets ---")
    for t in target_depts:
        print(t["label"])
    print("--- publishers ---")
    for a in accounts:
        if a["canPublishToOverall"]:
            print(
                f"{a['name']} / {a['phoneDept']}({a['phoneParent']}) / {a['phoneExt']}"
            )


if __name__ == "__main__":
    main()
